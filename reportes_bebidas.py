"""
reportes_bebidas.py - MÓDULO DE ANÁLISIS DE VENTAS CON EXPORTACIÓN A EXCEL

Dashboard profesional con estadísticas, comparativas y exportación a Excel
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from typing import List, Dict, Tuple
import sqlite3
import os

from database_bebidas import db
from ui_helpers_bebidas import (
    Tema,
    WidgetFactory,
    centrar_ventana,
    formatear_precio,
    formatear_numero,
)

# Importar openpyxl para Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_DISPONIBLE = True
except ImportError:
    EXCEL_DISPONIBLE = False


class VentanaReportesTealdi:
    """Dashboard de Análisis de Ventas - BEBIDAS TEALDI"""

    def __init__(self, sistema_principal):
        self.sistema = sistema_principal
        self.ventana = tk.Toplevel(self.sistema.root)
        self.ventana.title("📊 Reportes y Análisis - BEBIDAS TEALDI")

        # Detectar tamaño de pantalla
        screen_h = self.ventana.winfo_screenheight()
        self.es_pantalla_pequena = screen_h < 800

        if self.es_pantalla_pequena:
            ancho, alto = 1366, 768
        else:
            ancho, alto = 1400, 850

        centrar_ventana(self.ventana, ancho, alto)
        self.ventana.configure(bg=Tema.BG_MAIN)

        try:
            self.ventana.state("zoomed")
        except Exception:
            pass

        self.crear_interfaz()
        self.cargar_estadisticas()

    def crear_interfaz(self):
        pad = 10

        # Header
        self.crear_header()

        # Container principal con scroll
        main_container = tk.Frame(self.ventana, bg=Tema.BG_MAIN)
        main_container.pack(fill=tk.BOTH, expand=True, padx=pad, pady=pad)

        # Canvas con scroll
        canvas = tk.Canvas(main_container, bg=Tema.BG_MAIN, highlightthickness=0)
        scrollbar = ttk.Scrollbar(main_container, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg=Tema.BG_MAIN)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Secciones del dashboard
        self.crear_kpis_principales(scrollable_frame)
        self.crear_comparativas(scrollable_frame)
        self.crear_rankings(scrollable_frame)
        self.crear_productos_top(scrollable_frame)
        self.crear_ventas_por_categoria(scrollable_frame)
        self.crear_tabla_ventas_diarias(scrollable_frame)

        # NUEVO: Botón de exportación
        self.crear_boton_exportar(scrollable_frame)

    def crear_header(self):
        header_shadow = tk.Frame(self.ventana, bg=Tema.SHADOW, height=70)
        header_shadow.pack(fill=tk.X)
        header_shadow.pack_propagate(False)

        header_frame = tk.Frame(header_shadow, bg=Tema.PRIMARY)
        header_frame.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        left_frame = tk.Frame(header_frame, bg=Tema.PRIMARY)
        left_frame.pack(side=tk.LEFT, padx=20, pady=12)

        tk.Label(
            left_frame,
            text="📊",
            font=("Segoe UI Emoji", 28),
            bg=Tema.PRIMARY,
            fg=Tema.TEXT_LIGHT,
        ).pack(side=tk.LEFT, padx=(0, 12))

        tk.Label(
            left_frame,
            text="REPORTES Y ANÁLISIS DE VENTAS",
            font=(Tema.FONT_FAMILY, 22, "bold"),
            bg=Tema.PRIMARY,
            fg=Tema.TEXT_LIGHT,
        ).pack(side=tk.LEFT)

        right_frame = tk.Frame(header_frame, bg=Tema.PRIMARY)
        right_frame.pack(side=tk.RIGHT, padx=20, pady=12)

        tk.Label(
            right_frame,
            text=datetime.now().strftime("%B %Y").upper(),
            font=(Tema.FONT_FAMILY, 14, "bold"),
            bg=Tema.PRIMARY,
            fg=Tema.TEXT_LIGHT,
        ).pack()

    def crear_kpis_principales(self, parent):
        """KPIs principales en cards grandes"""
        kpis_frame = tk.Frame(parent, bg=Tema.BG_MAIN)
        kpis_frame.pack(fill=tk.X, pady=(0, 15))

        # Obtener datos
        hoy = datetime.now().strftime("%Y-%m-%d")
        mes_actual = datetime.now().strftime("%Y-%m")

        ventas_hoy = self.obtener_total_ventas_dia(hoy)
        ventas_mes = self.obtener_total_ventas_mes(mes_actual)
        cantidad_ventas_mes = self.obtener_cantidad_ventas_mes(mes_actual)
        ticket_promedio = ventas_mes / cantidad_ventas_mes if cantidad_ventas_mes > 0 else 0

        # Card 1: Ventas de hoy
        self.crear_kpi_card(
            kpis_frame, 
            "💰 VENTAS HOY",
            formatear_precio(ventas_hoy),
            Tema.SUCCESS,
            0
        )

        # Card 2: Ventas del mes
        self.crear_kpi_card(
            kpis_frame,
            "📈 VENTAS DEL MES",
            formatear_precio(ventas_mes),
            Tema.PRIMARY,
            1
        )

        # Card 3: Cantidad de ventas
        self.crear_kpi_card(
            kpis_frame,
            "🧾 TRANSACCIONES",
            formatear_numero(cantidad_ventas_mes),
            Tema.INFO,
            2
        )

        # Card 4: Ticket promedio
        self.crear_kpi_card(
            kpis_frame,
            "💳 TICKET PROMEDIO",
            formatear_precio(ticket_promedio),
            Tema.WARNING,
            3
        )

        kpis_frame.columnconfigure(0, weight=1)
        kpis_frame.columnconfigure(1, weight=1)
        kpis_frame.columnconfigure(2, weight=1)
        kpis_frame.columnconfigure(3, weight=1)

    def crear_kpi_card(self, parent, titulo: str, valor: str, color: str, col: int):
        shadow = tk.Frame(parent, bg=Tema.SHADOW)
        shadow.grid(row=0, column=col, sticky="ew", padx=5)

        card = tk.Frame(shadow, bg=color, height=120)
        card.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)
        card.pack_propagate(False)

        tk.Label(
            card,
            text=titulo,
            font=(Tema.FONT_FAMILY, 11, "bold"),
            bg=color,
            fg=Tema.TEXT_LIGHT,
        ).pack(pady=(20, 5))

        tk.Label(
            card,
            text=valor,
            font=(Tema.FONT_FAMILY, 24, "bold"),
            bg=color,
            fg=Tema.TEXT_LIGHT,
        ).pack()

    def crear_comparativas(self, parent):
        """Comparativas mes actual vs mes anterior"""
        shadow, comp_frame = WidgetFactory.crear_frame_card(parent)
        shadow.pack(fill=tk.X, pady=(0, 15), padx=5)

        tk.Label(
            comp_frame,
            text="📊 Comparativa Mensual",
            font=(Tema.FONT_FAMILY, 14, "bold"),
            bg=Tema.BG_CARD,
            fg=Tema.PRIMARY,
        ).pack(anchor=tk.W, padx=15, pady=(10, 5))

        # Obtener datos
        mes_actual = datetime.now().strftime("%Y-%m")
        mes_anterior = (datetime.now() - timedelta(days=30)).strftime("%Y-%m")

        ventas_actual = self.obtener_total_ventas_mes(mes_actual)
        ventas_anterior = self.obtener_total_ventas_mes(mes_anterior)

        diferencia = ventas_actual - ventas_anterior
        porcentaje = (diferencia / ventas_anterior * 100) if ventas_anterior > 0 else 0

        # Grid de comparación
        grid_frame = tk.Frame(comp_frame, bg=Tema.BG_CARD)
        grid_frame.pack(fill=tk.X, padx=15, pady=10)

        # Mes anterior
        tk.Label(
            grid_frame,
            text="Mes Anterior:",
            font=(Tema.FONT_FAMILY, 10),
            bg=Tema.BG_CARD,
            fg=Tema.TEXT_SECONDARY,
        ).grid(row=0, column=0, sticky="w", padx=(0, 20))

        tk.Label(
            grid_frame,
            text=formatear_precio(ventas_anterior),
            font=(Tema.FONT_FAMILY, 12, "bold"),
            bg=Tema.BG_CARD,
            fg=Tema.TEXT_PRIMARY,
        ).grid(row=0, column=1, sticky="w")

        # Mes actual
        tk.Label(
            grid_frame,
            text="Mes Actual:",
            font=(Tema.FONT_FAMILY, 10),
            bg=Tema.BG_CARD,
            fg=Tema.TEXT_SECONDARY,
        ).grid(row=1, column=0, sticky="w", padx=(0, 20), pady=(5, 0))

        tk.Label(
            grid_frame,
            text=formatear_precio(ventas_actual),
            font=(Tema.FONT_FAMILY, 12, "bold"),
            bg=Tema.BG_CARD,
            fg=Tema.TEXT_PRIMARY,
        ).grid(row=1, column=1, sticky="w", pady=(5, 0))

        # Diferencia
        color_diff = Tema.SUCCESS if diferencia >= 0 else Tema.DANGER
        simbolo = "▲" if diferencia >= 0 else "▼"

        tk.Label(
            grid_frame,
            text="Variación:",
            font=(Tema.FONT_FAMILY, 10),
            bg=Tema.BG_CARD,
            fg=Tema.TEXT_SECONDARY,
        ).grid(row=2, column=0, sticky="w", padx=(0, 20), pady=(5, 0))

        tk.Label(
            grid_frame,
            text=f"{simbolo} {formatear_precio(abs(diferencia))} ({abs(porcentaje):.1f}%)",
            font=(Tema.FONT_FAMILY, 12, "bold"),
            bg=Tema.BG_CARD,
            fg=color_diff,
        ).grid(row=2, column=1, sticky="w", pady=(5, 0))

    def crear_rankings(self, parent):
        """Rankings de mejores y peores días"""
        ranking_container = tk.Frame(parent, bg=Tema.BG_MAIN)
        ranking_container.pack(fill=tk.X, pady=(0, 15))

        # Mejor día
        shadow_mejor, mejor_frame = WidgetFactory.crear_frame_card(ranking_container)
        shadow_mejor.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 7.5))

        tk.Label(
            mejor_frame,
            text="🏆 MEJOR DÍA DEL MES",
            font=(Tema.FONT_FAMILY, 13, "bold"),
            bg=Tema.BG_CARD,
            fg=Tema.SUCCESS,
        ).pack(pady=(15, 10))

        mejor_dia = self.obtener_mejor_dia_mes()
        if mejor_dia:
            fecha, total = mejor_dia
            fecha_formateada = datetime.strptime(fecha, "%Y-%m-%d").strftime("%d/%m/%Y")

            tk.Label(
                mejor_frame,
                text=fecha_formateada,
                font=(Tema.FONT_FAMILY, 18, "bold"),
                bg=Tema.BG_CARD,
                fg=Tema.TEXT_PRIMARY,
            ).pack()

            tk.Label(
                mejor_frame,
                text=formatear_precio(total),
                font=(Tema.FONT_FAMILY, 22, "bold"),
                bg=Tema.BG_CARD,
                fg=Tema.SUCCESS,
            ).pack(pady=(5, 15))
        else:
            tk.Label(
                mejor_frame,
                text="Sin datos",
                font=(Tema.FONT_FAMILY, 14),
                bg=Tema.BG_CARD,
                fg=Tema.TEXT_SECONDARY,
            ).pack(pady=20)

        # Peor día
        shadow_peor, peor_frame = WidgetFactory.crear_frame_card(ranking_container)
        shadow_peor.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(7.5, 5))

        tk.Label(
            peor_frame,
            text="📉 PEOR DÍA DEL MES",
            font=(Tema.FONT_FAMILY, 13, "bold"),
            bg=Tema.BG_CARD,
            fg=Tema.DANGER,
        ).pack(pady=(15, 10))

        peor_dia = self.obtener_peor_dia_mes()
        if peor_dia:
            fecha, total = peor_dia
            fecha_formateada = datetime.strptime(fecha, "%Y-%m-%d").strftime("%d/%m/%Y")

            tk.Label(
                peor_frame,
                text=fecha_formateada,
                font=(Tema.FONT_FAMILY, 18, "bold"),
                bg=Tema.BG_CARD,
                fg=Tema.TEXT_PRIMARY,
            ).pack()

            tk.Label(
                peor_frame,
                text=formatear_precio(total),
                font=(Tema.FONT_FAMILY, 22, "bold"),
                bg=Tema.BG_CARD,
                fg=Tema.DANGER,
            ).pack(pady=(5, 15))
        else:
            tk.Label(
                peor_frame,
                text="Sin datos",
                font=(Tema.FONT_FAMILY, 14),
                bg=Tema.BG_CARD,
                fg=Tema.TEXT_SECONDARY,
            ).pack(pady=20)

    def crear_productos_top(self, parent):
        """Top 5 productos más vendidos"""
        shadow, top_frame = WidgetFactory.crear_frame_card(parent)
        shadow.pack(fill=tk.X, pady=(0, 15), padx=5)

        tk.Label(
            top_frame,
            text="⭐ TOP 5 PRODUCTOS MÁS VENDIDOS",
            font=(Tema.FONT_FAMILY, 14, "bold"),
            bg=Tema.BG_CARD,
            fg=Tema.PRIMARY,
        ).pack(anchor=tk.W, padx=15, pady=(10, 5))

        productos = self.obtener_productos_mas_vendidos(5)

        if productos:
            for i, (nombre, cantidad, total) in enumerate(productos, 1):
                item_frame = tk.Frame(top_frame, bg=Tema.BG_CARD)
                item_frame.pack(fill=tk.X, padx=15, pady=5)

                # Medalla según posición
                medallas = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]

                tk.Label(
                    item_frame,
                    text=medallas[i-1],
                    font=(Tema.FONT_FAMILY, 16),
                    bg=Tema.BG_CARD,
                ).pack(side=tk.LEFT, padx=(0, 10))

                tk.Label(
                    item_frame,
                    text=nombre,
                    font=(Tema.FONT_FAMILY, 11, "bold"),
                    bg=Tema.BG_CARD,
                    fg=Tema.TEXT_PRIMARY,
                    anchor=tk.W,
                ).pack(side=tk.LEFT, fill=tk.X, expand=True)

                tk.Label(
                    item_frame,
                    text=f"{cantidad} unid.",
                    font=(Tema.FONT_FAMILY, 10),
                    bg=Tema.BG_CARD,
                    fg=Tema.TEXT_SECONDARY,
                ).pack(side=tk.LEFT, padx=10)

                tk.Label(
                    item_frame,
                    text=formatear_precio(total),
                    font=(Tema.FONT_FAMILY, 11, "bold"),
                    bg=Tema.BG_CARD,
                    fg=Tema.SUCCESS,
                ).pack(side=tk.LEFT)
        else:
            tk.Label(
                top_frame,
                text="No hay datos de productos vendidos",
                font=(Tema.FONT_FAMILY, 11),
                bg=Tema.BG_CARD,
                fg=Tema.TEXT_SECONDARY,
            ).pack(pady=20)

    def crear_ventas_por_categoria(self, parent):
        """Análisis por categoría"""
        shadow, cat_frame = WidgetFactory.crear_frame_card(parent)
        shadow.pack(fill=tk.X, pady=(0, 15), padx=5)

        tk.Label(
            cat_frame,
            text="📦 VENTAS POR CATEGORÍA",
            font=(Tema.FONT_FAMILY, 14, "bold"),
            bg=Tema.BG_CARD,
            fg=Tema.PRIMARY,
        ).pack(anchor=tk.W, padx=15, pady=(10, 5))

        categorias = self.obtener_ventas_por_categoria()

        if categorias:
            for categoria, cantidad, total in categorias:
                item_frame = tk.Frame(cat_frame, bg=Tema.BG_CARD)
                item_frame.pack(fill=tk.X, padx=15, pady=5)

                tk.Label(
                    item_frame,
                    text=categoria,
                    font=(Tema.FONT_FAMILY, 11, "bold"),
                    bg=Tema.BG_CARD,
                    fg=Tema.TEXT_PRIMARY,
                    anchor=tk.W,
                    width=15,
                ).pack(side=tk.LEFT)

                # Barra de progreso visual
                total_max = categorias[0][2]  # El más alto
                porcentaje = (total / total_max * 100) if total_max > 0 else 0

                barra_frame = tk.Frame(item_frame, bg=Tema.BG_MAIN, height=20)
                barra_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)

                barra_inner = tk.Frame(barra_frame, bg=Tema.PRIMARY, height=20)
                barra_inner.place(relwidth=porcentaje/100, relheight=1)

                tk.Label(
                    item_frame,
                    text=f"{cantidad} un.",
                    font=(Tema.FONT_FAMILY, 10),
                    bg=Tema.BG_CARD,
                    fg=Tema.TEXT_SECONDARY,
                    width=10,
                ).pack(side=tk.LEFT)

                tk.Label(
                    item_frame,
                    text=formatear_precio(total),
                    font=(Tema.FONT_FAMILY, 11, "bold"),
                    bg=Tema.BG_CARD,
                    fg=Tema.SUCCESS,
                    width=12,
                    anchor=tk.E,
                ).pack(side=tk.LEFT)
        else:
            tk.Label(
                cat_frame,
                text="No hay datos de categorías",
                font=(Tema.FONT_FAMILY, 11),
                bg=Tema.BG_CARD,
                fg=Tema.TEXT_SECONDARY,
            ).pack(pady=20)

    def crear_tabla_ventas_diarias(self, parent):
        """Tabla con ventas diarias del mes"""
        shadow, tabla_frame = WidgetFactory.crear_frame_card(parent)
        shadow.pack(fill=tk.BOTH, expand=True, pady=(0, 15), padx=5)

        tk.Label(
            tabla_frame,
            text="📅 VENTAS DIARIAS DEL MES",
            font=(Tema.FONT_FAMILY, 14, "bold"),
            bg=Tema.BG_CARD,
            fg=Tema.PRIMARY,
        ).pack(anchor=tk.W, padx=15, pady=(10, 5))

        # Tabla
        tree_container = tk.Frame(tabla_frame, bg=Tema.SHADOW)
        tree_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=(0, 15))

        tree_inner = tk.Frame(tree_container, bg=Tema.BG_CARD)
        tree_inner.pack(padx=2, pady=2, fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(tree_inner, orient=tk.VERTICAL)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        tree = ttk.Treeview(
            tree_inner,
            columns=("fecha", "dia", "ventas", "total", "ticket_prom"),
            show="headings",
            yscrollcommand=scrollbar.set,
            height=10,
            style="Modern.Treeview",
        )

        tree.heading("fecha", text="Fecha")
        tree.heading("dia", text="Día")
        tree.heading("ventas", text="N° Ventas")
        tree.heading("total", text="Total Vendido")
        tree.heading("ticket_prom", text="Ticket Prom.")

        tree.column("fecha", width=100, anchor=tk.CENTER)
        tree.column("dia", width=100, anchor=tk.CENTER)
        tree.column("ventas", width=100, anchor=tk.CENTER)
        tree.column("total", width=150, anchor=tk.CENTER)
        tree.column("ticket_prom", width=150, anchor=tk.CENTER)

        tree.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)
        scrollbar.config(command=tree.yview)

        # Cargar datos
        ventas_diarias = self.obtener_ventas_diarias_mes()

        for fecha, cantidad, total in ventas_diarias:
            fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")
            fecha_format = fecha_obj.strftime("%d/%m/%Y")
            dia_semana = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][fecha_obj.weekday()]
            ticket_prom = total / cantidad if cantidad > 0 else 0

            tree.insert("", tk.END, values=(
                fecha_format,
                dia_semana,
                cantidad,
                formatear_precio(total),
                formatear_precio(ticket_prom)
            ))

    def crear_boton_exportar(self, parent):
        """NUEVO: Botón para exportar reporte a Excel"""
        shadow, export_frame = WidgetFactory.crear_frame_card(parent)
        shadow.pack(fill=tk.X, pady=(0, 15), padx=5)

        tk.Label(
            export_frame,
            text="📥 EXPORTAR REPORTE",
            font=(Tema.FONT_FAMILY, 14, "bold"),
            bg=Tema.BG_CARD,
            fg=Tema.PRIMARY,
        ).pack(anchor=tk.W, padx=15, pady=(10, 5))

        info_frame = tk.Frame(export_frame, bg=Tema.BG_CARD)
        info_frame.pack(fill=tk.X, padx=15, pady=(0, 10))

        tk.Label(
            info_frame,
            text="Exporta todos los reportes y estadísticas a un archivo Excel profesional",
            font=(Tema.FONT_FAMILY, 10),
            bg=Tema.BG_CARD,
            fg=Tema.TEXT_SECONDARY,
        ).pack(anchor=tk.W)

        if not EXCEL_DISPONIBLE:
            tk.Label(
                export_frame,
                text="⚠️ Instala openpyxl para exportar: pip install openpyxl",
                font=(Tema.FONT_FAMILY, 10),
                bg=Tema.BG_CARD,
                fg=Tema.DANGER,
            ).pack(padx=15, pady=10)
        else:
            WidgetFactory.crear_boton(
                export_frame,
                "📥 EXPORTAR A EXCEL",
                self.exportar_a_excel,
                "success",
                width=32
            ).pack(padx=15, pady=(0, 15))

    # ============ EXPORTACIÓN A EXCEL ============

    def exportar_a_excel(self):
        """Exporta el reporte completo a Excel con formato profesional"""
        if not EXCEL_DISPONIBLE:
            messagebox.showerror(
                "Error - Tealdi",
                "Librería openpyxl no instalada.\n\nInstala con: pip install openpyxl",
                parent=self.ventana
            )
            return

        try:
            # Pedir ubicación para guardar
            mes_actual = datetime.now().strftime("%Y-%m")
            nombre_archivo = f"Reporte_Tealdi_{mes_actual}.xlsx"

            archivo = filedialog.asksaveasfilename(
                parent=self.ventana,
                title="Guardar Reporte Excel",
                defaultextension=".xlsx",
                initialfile=nombre_archivo,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )

            if not archivo:
                return

            # Crear workbook
            wb = Workbook()

            # Crear hojas
            self._crear_hoja_resumen(wb)
            self._crear_hoja_ventas_diarias(wb)
            self._crear_hoja_productos_top(wb)
            self._crear_hoja_categorias(wb)

            # Eliminar hoja por defecto
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Guardar archivo
            wb.save(archivo)

            # Mostrar mensaje de éxito
            messagebox.showinfo(
                "Exportación Exitosa - Tealdi",
                f"Reporte exportado exitosamente en:\n{archivo}",
                parent=self.ventana
            )

            # Abrir archivo
            os.startfile(archivo)

        except Exception as e:
            messagebox.showerror(
                "Error - Tealdi",
                f"Error al exportar reporte:\n{str(e)}",
                parent=self.ventana
            )

    def _crear_hoja_resumen(self, wb):
        """Crea hoja de resumen con KPIs principales"""
        ws = wb.active
        ws.title = "Resumen"

        # Estilos
        header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título
        ws['A1'] = "REPORTE DE VENTAS - BEBIDAS TEALDI"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')

        ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws.merge_cells('A2:D2')

        # KPIs
        row = 4
        ws[f'A{row}'] = "INDICADORES CLAVE"
        ws[f'A{row}'].font = Font(bold=True, size=14)

        row += 2
        mes_actual = datetime.now().strftime("%Y-%m")
        hoy = datetime.now().strftime("%Y-%m-%d")

        ventas_hoy = self.obtener_total_ventas_dia(hoy)
        ventas_mes = self.obtener_total_ventas_mes(mes_actual)
        cantidad_ventas = self.obtener_cantidad_ventas_mes(mes_actual)
        ticket_promedio = ventas_mes / cantidad_ventas if cantidad_ventas > 0 else 0

        kpis = [
            ("💰 Ventas de Hoy", f"${ventas_hoy:,.2f}"),
            ("📈 Ventas del Mes", f"${ventas_mes:,.2f}"),
            ("🧾 Cantidad de Transacciones", cantidad_ventas),
            ("💳 Ticket Promedio", f"${ticket_promedio:,.2f}"),
        ]

        for kpi_name, kpi_value in kpis:
            ws[f'A{row}'] = kpi_name
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = kpi_value
            ws[f'B{row}'].font = Font(size=12)
            row += 1

        # Comparativa
        row += 2
        ws[f'A{row}'] = "COMPARATIVA MENSUAL"
        ws[f'A{row}'].font = Font(bold=True, size=14)

        row += 2
        mes_anterior = (datetime.now() - timedelta(days=30)).strftime("%Y-%m")
        ventas_anterior = self.obtener_total_ventas_mes(mes_anterior)
        diferencia = ventas_mes - ventas_anterior

        ws[f'A{row}'] = "Mes Anterior"
        ws[f'B{row}'] = f"${ventas_anterior:,.2f}"
        row += 1
        ws[f'A{row}'] = "Mes Actual"
        ws[f'B{row}'] = f"${ventas_mes:,.2f}"
        row += 1
        ws[f'A{row}'] = "Diferencia"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"${diferencia:,.2f}"
        ws[f'B{row}'].font = Font(bold=True, color="00B050" if diferencia >= 0 else "FF0000")

        # Ajustar anchos
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20

    def _crear_hoja_ventas_diarias(self, wb):
        """Crea hoja con ventas diarias"""
        ws = wb.create_sheet("Ventas Diarias")

        # Headers
        headers = ["Fecha", "Día", "N° Ventas", "Total Vendido", "Ticket Promedio"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Datos
        ventas_diarias = self.obtener_ventas_diarias_mes()
        row = 2

        for fecha, cantidad, total in ventas_diarias:
            fecha_obj = datetime.strptime(fecha, "%Y-%m-%d")
            fecha_format = fecha_obj.strftime("%d/%m/%Y")
            dia_semana = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][fecha_obj.weekday()]
            ticket_prom = total / cantidad if cantidad > 0 else 0

            ws.cell(row, 1, fecha_format)
            ws.cell(row, 2, dia_semana)
            ws.cell(row, 3, cantidad)
            ws.cell(row, 4, f"${total:,.2f}")
            ws.cell(row, 5, f"${ticket_prom:,.2f}")
            row += 1

        # Totales
        total_ventas = sum(v[1] for v in ventas_diarias)
        total_monto = sum(v[2] for v in ventas_diarias)

        ws.cell(row, 1, "TOTALES").font = Font(bold=True)
        ws.cell(row, 3, total_ventas).font = Font(bold=True)
        ws.cell(row, 4, f"${total_monto:,.2f}").font = Font(bold=True)

        # Ajustar anchos
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 18

    def _crear_hoja_productos_top(self, wb):
        """Crea hoja con productos más vendidos"""
        ws = wb.create_sheet("Top Productos")

        # Headers
        headers = ["#", "Producto", "Cantidad Vendida", "Total Generado"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Datos
        productos = self.obtener_productos_mas_vendidos(20)
        row = 2

        for i, (nombre, cantidad, total) in enumerate(productos, 1):
            ws.cell(row, 1, i)
            ws.cell(row, 2, nombre)
            ws.cell(row, 3, cantidad)
            ws.cell(row, 4, f"${total:,.2f}")
            row += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18

    def _crear_hoja_categorias(self, wb):
        """Crea hoja con ventas por categoría"""
        ws = wb.create_sheet("Por Categoría")

        # Headers
        headers = ["Categoría", "Cantidad Vendida", "Total Generado", "% del Total"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Datos
        categorias = self.obtener_ventas_por_categoria()
        total_general = sum(c[2] for c in categorias)
        row = 2

        for categoria, cantidad, total in categorias:
            porcentaje = (total / total_general * 100) if total_general > 0 else 0
            ws.cell(row, 1, categoria)
            ws.cell(row, 2, cantidad)
            ws.cell(row, 3, f"${total:,.2f}")
            ws.cell(row, 4, f"{porcentaje:.1f}%")
            row += 1

        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 15

    # ============ MÉTODOS DE DATOS (sin cambios) ============

    def obtener_total_ventas_dia(self, fecha: str) -> float:
        try:
            conn = sqlite3.connect("bebidas_tealdi.db")
            cursor = conn.cursor()
            cursor.execute("SELECT SUM(total) FROM ventas WHERE DATE(fecha) = ?", (fecha,))
            resultado = cursor.fetchone()[0]
            conn.close()
            return resultado if resultado else 0.0
        except Exception:
            return 0.0

    def obtener_total_ventas_mes(self, mes: str) -> float:
        try:
            conn = sqlite3.connect("bebidas_tealdi.db")
            cursor = conn.cursor()
            cursor.execute("SELECT SUM(total) FROM ventas WHERE strftime('%Y-%m', fecha) = ?", (mes,))
            resultado = cursor.fetchone()[0]
            conn.close()
            return resultado if resultado else 0.0
        except Exception:
            return 0.0

    def obtener_cantidad_ventas_mes(self, mes: str) -> int:
        try:
            conn = sqlite3.connect("bebidas_tealdi.db")
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM ventas WHERE strftime('%Y-%m', fecha) = ?", (mes,))
            resultado = cursor.fetchone()[0]
            conn.close()
            return resultado if resultado else 0
        except Exception:
            return 0

    def obtener_mejor_dia_mes(self) -> Tuple[str, float]:
        try:
            mes_actual = datetime.now().strftime("%Y-%m")
            conn = sqlite3.connect("bebidas_tealdi.db")
            cursor = conn.cursor()
            cursor.execute("""
                SELECT DATE(fecha), SUM(total) as total_dia
                FROM ventas
                WHERE strftime('%Y-%m', fecha) = ?
                GROUP BY DATE(fecha)
                ORDER BY total_dia DESC
                LIMIT 1
            """, (mes_actual,))
            resultado = cursor.fetchone()
            conn.close()
            return resultado if resultado else None
        except Exception:
            return None

    def obtener_peor_dia_mes(self) -> Tuple[str, float]:
        try:
            mes_actual = datetime.now().strftime("%Y-%m")
            conn = sqlite3.connect("bebidas_tealdi.db")
            cursor = conn.cursor()
            cursor.execute("""
                SELECT DATE(fecha), SUM(total) as total_dia
                FROM ventas
                WHERE strftime('%Y-%m', fecha) = ?
                GROUP BY DATE(fecha)
                ORDER BY total_dia ASC
                LIMIT 1
            """, (mes_actual,))
            resultado = cursor.fetchone()
            conn.close()
            return resultado if resultado else None
        except Exception:
            return None

    def obtener_productos_mas_vendidos(self, limite: int = 5) -> List[Tuple]:
        try:
            mes_actual = datetime.now().strftime("%Y-%m")
            conn = sqlite3.connect("bebidas_tealdi.db")
            cursor = conn.cursor()
            cursor.execute("""
                SELECT vi.nombre, SUM(vi.cantidad) as total_cant, SUM(vi.subtotal) as total_venta
                FROM ventas_items vi
                JOIN ventas v ON vi.venta_id = v.id
                WHERE strftime('%Y-%m', v.fecha) = ?
                GROUP BY vi.nombre
                ORDER BY total_cant DESC
                LIMIT ?
            """, (mes_actual, limite))
            resultados = cursor.fetchall()
            conn.close()
            return resultados
        except Exception:
            return []

    def obtener_ventas_por_categoria(self) -> List[Tuple]:
        try:
            mes_actual = datetime.now().strftime("%Y-%m")
            conn = sqlite3.connect("bebidas_tealdi.db")
            cursor = conn.cursor()
            cursor.execute("""
                SELECT vi.categoria, SUM(vi.cantidad) as total_cant, SUM(vi.subtotal) as total_venta
                FROM ventas_items vi
                JOIN ventas v ON vi.venta_id = v.id
                WHERE strftime('%Y-%m', v.fecha) = ?
                GROUP BY vi.categoria
                ORDER BY total_venta DESC
            """, (mes_actual,))
            resultados = cursor.fetchall()
            conn.close()
            return resultados
        except Exception:
            return []

    def obtener_ventas_diarias_mes(self) -> List[Tuple]:
        try:
            mes_actual = datetime.now().strftime("%Y-%m")
            conn = sqlite3.connect("bebidas_tealdi.db")
            cursor = conn.cursor()
            cursor.execute("""
                SELECT DATE(fecha), COUNT(*) as num_ventas, SUM(total) as total_dia
                FROM ventas
                WHERE strftime('%Y-%m', fecha) = ?
                GROUP BY DATE(fecha)
                ORDER BY DATE(fecha) DESC
            """, (mes_actual,))
            resultados = cursor.fetchall()
            conn.close()
            return resultados
        except Exception:
            return []

    def cargar_estadisticas(self):
        """Carga todas las estadísticas al abrir la ventana"""
        pass


def abrir_ventana_reportes_bebidas(sistema_principal):
    """Función para abrir la ventana de reportes desde el main"""
    VentanaReportesTealdi(sistema_principal)


if __name__ == "__main__":
    root = tk.Tk()
    root.withdraw()
    class FakeSistema:
        def __init__(self, root_):
            self.root = root_
    VentanaReportesTealdi(FakeSistema(root))
    root.mainloop()
